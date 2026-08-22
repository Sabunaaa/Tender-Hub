from __future__ import annotations

import io

from tender_scraper.specs import (
    CELL_SEP,
    FILE_PREFIX,
    SHEET_PREFIX,
    extract_bytes,
    extract_spec_text,
    is_spec_filename,
    spec_files,
)


def test_is_spec_filename_requires_georgian_marker():
    assert is_spec_filename("ტექნიკური დავალება.xlsx")
    assert is_spec_filename("დანართი N1 - ტექნიკური აღწერილობა.pdf")
    assert not is_spec_filename("Technical specification.xlsx")
    assert not is_spec_filename("ხელშეკრულების პროექტი.pdf")
    assert not is_spec_filename("")


def test_spec_files_filters_and_caps():
    atts = [
        {"name": "ტექნიკური დავალება.xlsx", "url": "https://example.com/a.xlsx"},
        {"name": "ტექნიკური დავალება.xlsx", "url": "https://example.com/a.xlsx"},
        {"name": "ხელშეკრულება.pdf", "url": "https://example.com/b.pdf"},
        {"name": "ტექნიკური.doc", "url": "https://example.com/old.doc"},
        {"name": "ტექნიკური აღწერილობა.pdf", "url": "https://example.com/c.pdf"},
    ]
    files = spec_files(atts)
    assert [f["url"] for f in files] == [
        "https://example.com/a.xlsx",
        "https://example.com/c.pdf",
    ]


def test_extract_xlsx_cells():
    from openpyxl import Workbook

    wb = Workbook()
    sheet = wb.active
    sheet["A1"] = "კომუტატორი"
    sheet["B1"] = "24 პორტი"
    buf = io.BytesIO()
    wb.save(buf)
    text = extract_bytes(buf.getvalue(), "ტექნიკური დავალება.xlsx")
    assert "კომუტატორი" in text
    assert "24 პორტი" in text


def test_extract_spec_text_skips_non_matching(monkeypatch):
    class FakeClient:
        def get_bytes(self, url: str) -> bytes:
            raise AssertionError(f"should not download {url}")

    text = extract_spec_text(
        FakeClient(),
        [{"name": "ხელშეკრულება.pdf", "url": "https://example.com/x.pdf"}],
    )
    assert text == ""


def test_extract_spec_text_uses_matching_xlsx():
    from openpyxl import Workbook

    wb = Workbook()
    wb.active["A1"] = "მარშრუტიზატორი"
    buf = io.BytesIO()
    wb.save(buf)
    payload = buf.getvalue()

    class FakeClient:
        def get_bytes(self, url: str) -> bytes:
            assert url == "https://example.com/spec.xlsx"
            return payload

    text = extract_spec_text(
        FakeClient(),
        [
            {"name": "ტექნიკური დავალება.xlsx", "url": "https://example.com/spec.xlsx"},
            {"name": "ფასები.xlsx", "url": "https://example.com/prices.xlsx"},
        ],
    )
    assert "მარშრუტიზატორი" in text


def test_xlsx_keeps_rows_and_cells():
    """The spec panel rebuilds tables from these separators, so they must survive."""
    from openpyxl import Workbook

    wb = Workbook()
    sheet = wb.active
    sheet.append(["N", "დასახელება", "რაოდენობა"])
    sheet.append([1, "კომუტატორი", 20])
    sheet.append([None, None, None])
    sheet.append([2, "მარშრუტიზატორი", 5])
    buf = io.BytesIO()
    wb.save(buf)

    lines = extract_bytes(buf.getvalue(), "ტექნიკური დავალება.xlsx").split("\n")

    assert lines == [
        f"N{CELL_SEP}დასახელება{CELL_SEP}რაოდენობა",
        f"1{CELL_SEP}კომუტატორი{CELL_SEP}20",
        f"2{CELL_SEP}მარშრუტიზატორი{CELL_SEP}5",
    ]


def test_xlsx_labels_each_sheet_when_there_are_several():
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "კომუტატორები"
    wb.active["A1"] = "24 პორტი"
    second = wb.create_sheet("როუტერები")
    second["A1"] = "SD-WAN"
    buf = io.BytesIO()
    wb.save(buf)

    text = extract_bytes(buf.getvalue(), "ტექნიკური დავალება.xlsx")

    assert f"{SHEET_PREFIX}კომუტატორები" in text
    assert f"{SHEET_PREFIX}როუტერები" in text


def test_spec_text_labels_each_source_file():
    from openpyxl import Workbook

    wb = Workbook()
    wb.active["A1"] = "კომუტატორი"
    buf = io.BytesIO()
    wb.save(buf)
    payload = buf.getvalue()

    class FakeClient:
        def get_bytes(self, url: str) -> bytes:
            return payload

    text = extract_spec_text(
        FakeClient(),
        [{"name": "ტექნიკური დავალება.xlsx", "url": "https://example.com/spec.xlsx"}],
    )

    assert text.startswith(f"{FILE_PREFIX}ტექნიკური დავალება.xlsx\n")


def test_extract_pdf_does_not_raise():
    from pypdf import PdfWriter

    writer = PdfWriter()
    writer.add_blank_page(width=200, height=200)
    buf = io.BytesIO()
    writer.write(buf)
    text = extract_bytes(buf.getvalue(), "ტექნიკური.pdf")
    assert isinstance(text, str)
